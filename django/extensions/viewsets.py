from rest_framework.mixins import ListModelMixin, CreateModelMixin, RetrieveModelMixin, UpdateModelMixin, DestroyModelMixin
from extensions.paginations import LimitedPagination, InfinitePagination
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models.deletion import ProtectedError
from rest_framework.viewsets import GenericViewSet
from extensions.exceptions import ValidationError
from django.http.response import HttpResponse
from openpyxl import Workbook, load_workbook
from rest_framework.viewsets import ViewSet


class FunctionViewSet(ViewSet):
    """機能ビュー"""

    @property
    def team(self):
        return self.request.user.team

    @property
    def user(self):
        return self.request.user

    @property
    def context(self):
        return {'request': self.request, 'format': self.format_kwarg, 'view': self}


class BaseViewSet(GenericViewSet):
    pagination_class = LimitedPagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    ordering_fields = ['id']
    ordering = ['-id']
    select_related_fields = []
    prefetch_related_fields = []

    @property
    def team(self):
        return self.request.user.team

    @property
    def user(self):
        return self.request.user

    @property
    def context(self):
        return self.get_serializer_context()

    def get_queryset(self):
        queryset = super().get_queryset().filter(team=self.team)
        queryset = queryset.select_related(*self.select_related_fields)
        queryset = queryset.prefetch_related(*self.prefetch_related_fields)
        return queryset
    
    
class SafeDestroyMixin(DestroyModelMixin):
    
    def perform_destroy(self, instance):
        try:
            instance.delete()
        except ProtectedError:
            raise ValidationError('データから参照されているため削除できません')


class ModelViewSet(BaseViewSet, ListModelMixin, CreateModelMixin,
                   RetrieveModelMixin, UpdateModelMixin, SafeDestroyMixin):
    """モデルビュー"""


class PersonalViewSet(BaseViewSet):
    """個人ビュー"""

    def get_queryset(self):
        return super().get_queryset().filter(creator=self.user)


class LimitedOptionViewSet(BaseViewSet, ListModelMixin):
    """有限オプションビュー"""


class InfiniteOptionViewSet(BaseViewSet, ListModelMixin):
    """無限オプションビュー"""

    pagination_class = InfinitePagination


class ReadOnlyMixin(RetrieveModelMixin, ListModelMixin):
    """読み取り専用"""


class ExportMixin:
    """エクスポート"""

    def get_export_response(self, serializer_class, data=None):
        """エクスポートExcelファイルレスポンスを取得

        Args:
            serializer_class (BaseSerializer): シリアライザークラス
            data (list): データ

            serializer (BaseSerializer): シリアライザー
            field_items (List): フィールド属性

        Returns:
            HttpResponse: ファイルレスポンス
        """

        workbook = Workbook()
        work_sheet = workbook.active

        if data:
            results = data
        else:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = serializer_class(instance=queryset, many=True)
            results = serializer.data
        field_items = serializer_class().get_fields().items()

        # テーブルヘッダーを作成
        for index, (field_name, field_class) in enumerate(field_items, start=1):
            work_sheet.cell(row=1, column=index, value=field_class.label)

        # データを入力
        for row, item in enumerate(results, start=2):
            for column, (field_name, field_class) in enumerate(field_items, start=1):
                work_sheet.cell(row=row, column=column, value=item.get(field_name, ''))

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=data.xlsx'
        workbook.save(response)

        return response


class ImportMixin:
    """インポート"""

    def get_template_response(self, serializer_class):
        """Excelテンプレートレスポンスを取得

        Args:
            serializer_class (BaseSerializer): シリアライザークラス

        Returns:
            HttpResponse: ファイルレスポンス
        """

        workbook = Workbook()
        work_sheet = workbook.active

        # テーブルヘッダーを作成
        field_items = serializer_class().get_fields().items()
        for index, (_, field_class) in enumerate(field_items, start=1):
            work_sheet.cell(row=1, column=index, value=field_class.label)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=data.xlsx'
        workbook.save(response)

        return response

    def load_data(self, file, serializer_class):
        workbook = load_workbook(file)
        work_sheet = workbook.active
        field_items = serializer_class().get_fields().items()

        for column, (field_name, field_class) in enumerate(field_items):
            if work_sheet[1][column].value != field_class.label:
                raise ValidationError('フォーマットエラー')

        data = []
        for row in range(2, work_sheet.max_row + 1):
            instance_item = {}
            for column, (field_name, field_class) in enumerate(field_items):
                if work_sheet[row][column].value is not None:
                    instance_item[field_name] = work_sheet[row][column].value
            else:
                data.append(instance_item)

        serializer = serializer_class(data=data, many=True, context=self.context)
        return serializer


__all__ = [
    'FunctionViewSet', 'BaseViewSet', 'ModelViewSet', 'PersonalViewSet',
    'LimitedOptionViewSet', 'InfiniteOptionViewSet',
    'ReadOnlyMixin', 'SafeDestroyMixin', 'ExportMixin', 'ImportMixin',
    'ListModelMixin', 'CreateModelMixin', 'RetrieveModelMixin', 'UpdateModelMixin', 'DestroyModelMixin',
    'DjangoFilterBackend', 'SearchFilter', 'OrderingFilter',
]
